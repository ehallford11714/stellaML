"""Tabular auto-structure, cleaning, EDA, charting, and insight helpers."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, median, stdev


@dataclass(slots=True)
class EDAReport:
    rows: int
    columns: list[str]
    inferred_types: dict[str, str]
    missing_counts: dict[str, int]
    numeric_summary: dict[str, dict[str, float]]
    categorical_summary: dict[str, dict[str, int]]
    recommendations: list[str]


def load_and_impute_csv(csv_path: str | Path) -> tuple[list[dict[str, object]], dict[str, str]]:
    """Backwards-compatible CSV loader with auto-type inference + imputation."""
    rows = _read_rows(csv_path)
    inferred_types = infer_structure(rows)
    typed_rows = cast_rows(rows, inferred_types)
    impute_missing(typed_rows, inferred_types)
    return typed_rows, inferred_types


def load_tabular_file(path: str | Path) -> tuple[list[dict[str, object]], dict[str, str]]:
    """Load CSV or Excel, infer structure, cast values, and impute missing data."""
    rows = _read_rows(path)
    inferred_types = infer_structure(rows)
    typed_rows = cast_rows(rows, inferred_types)
    impute_missing(typed_rows, inferred_types)
    return typed_rows, inferred_types


def infer_structure(rows: list[dict[str, object]]) -> dict[str, str]:
    if not rows:
        return {}

    columns = list(rows[0].keys())
    inferred_types: dict[str, str] = {}
    for col in columns:
        non_empty = [str(r[col]).strip() for r in rows if r.get(col) not in (None, "")]
        numeric_ratio = (sum(_is_float(v) for v in non_empty) / len(non_empty)) if non_empty else 0.0
        inferred_types[col] = "numeric" if numeric_ratio >= 0.7 else "categorical"
    return inferred_types


def cast_rows(rows: list[dict[str, object]], inferred_types: dict[str, str]) -> list[dict[str, object]]:
    typed_rows: list[dict[str, object]] = []
    for row in rows:
        out: dict[str, object] = {}
        for col, col_type in inferred_types.items():
            raw = row.get(col)
            value = str(raw).strip() if raw is not None else ""
            if value == "":
                out[col] = None
            elif col_type == "numeric":
                out[col] = float(value) if _is_float(value) else None
            else:
                out[col] = value
        typed_rows.append(out)
    return typed_rows


def impute_missing(rows: list[dict[str, object]], inferred_types: dict[str, str]) -> None:
    if not rows:
        return

    for col, col_type in inferred_types.items():
        values = [r[col] for r in rows if r.get(col) is not None]
        if not values:
            continue

        if col_type == "numeric":
            fill = mean(float(v) for v in values)
        else:
            fill = Counter(str(v) for v in values).most_common(1)[0][0]

        for row in rows:
            if row.get(col) is None:
                row[col] = fill


def apply_cleaning_operations(
    rows: list[dict[str, object]],
    inferred_types: dict[str, str],
    operations: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Apply cleaning ops such as binarization, discretization, encoding, etc."""
    cleaned = [dict(row) for row in rows]

    for op in operations:
        name = str(op.get("op", "")).lower()
        column = str(op.get("column", ""))

        if name == "drop_duplicates":
            signature_seen: set[tuple[tuple[str, object], ...]] = set()
            deduped: list[dict[str, object]] = []
            for row in cleaned:
                key = tuple(sorted(row.items()))
                if key not in signature_seen:
                    signature_seen.add(key)
                    deduped.append(row)
            cleaned = deduped

        elif name == "binarize":
            threshold = float(op.get("threshold", 0.0))
            new_column = str(op.get("new_column", f"{column}_bin"))
            for row in cleaned:
                value = row.get(column)
                row[new_column] = 1 if isinstance(value, (float, int)) and float(value) >= threshold else 0
            inferred_types[new_column] = "numeric"

        elif name == "discretize":
            bins = int(op.get("bins", 4))
            new_column = str(op.get("new_column", f"{column}_bucket"))
            values = [float(r[column]) for r in cleaned if isinstance(r.get(column), (float, int))]
            if values and bins > 0:
                lo, hi = min(values), max(values)
                width = (hi - lo) / bins if hi != lo else 1.0
                for row in cleaned:
                    value = row.get(column)
                    if isinstance(value, (float, int)):
                        idx = int(min((float(value) - lo) // width, bins - 1)) if width else 0
                        row[new_column] = f"bin_{idx}"
                    else:
                        row[new_column] = "unknown"
                inferred_types[new_column] = "categorical"

        elif name == "normalize_minmax":
            values = [float(r[column]) for r in cleaned if isinstance(r.get(column), (float, int))]
            if values:
                lo, hi = min(values), max(values)
                rng = hi - lo if hi != lo else 1.0
                for row in cleaned:
                    value = row.get(column)
                    if isinstance(value, (float, int)):
                        row[column] = (float(value) - lo) / rng

        elif name == "standardize":
            values = [float(r[column]) for r in cleaned if isinstance(r.get(column), (float, int))]
            if len(values) > 1:
                mu = mean(values)
                sigma = stdev(values) or 1.0
                for row in cleaned:
                    value = row.get(column)
                    if isinstance(value, (float, int)):
                        row[column] = (float(value) - mu) / sigma

        elif name == "one_hot_encode":
            categories = sorted({str(r[column]) for r in cleaned if r.get(column) is not None})
            for cat in categories:
                new_col = f"{column}__{cat}"
                for row in cleaned:
                    row[new_col] = 1 if str(row.get(column)) == cat else 0
                inferred_types[new_col] = "numeric"

        elif name == "fill_missing":
            strategy = str(op.get("strategy", "mean")).lower()
            values = [r[column] for r in cleaned if r.get(column) is not None]
            if not values:
                continue
            if strategy == "median" and inferred_types.get(column) == "numeric":
                fill = median(float(v) for v in values)
            elif strategy == "mode":
                fill = Counter(str(v) for v in values).most_common(1)[0][0]
            else:
                fill = mean(float(v) for v in values) if inferred_types.get(column) == "numeric" else Counter(str(v) for v in values).most_common(1)[0][0]
            for row in cleaned:
                if row.get(column) is None:
                    row[column] = fill

    return cleaned


def auto_eda(rows: list[dict[str, object]], inferred_types: dict[str, str]) -> EDAReport:
    if not rows:
        return EDAReport(0, [], {}, {}, {}, {}, ["Load a non-empty dataset to run EDA."])

    columns = list(rows[0].keys())
    missing_counts: dict[str, int] = {}
    numeric_summary: dict[str, dict[str, float]] = {}
    categorical_summary: dict[str, dict[str, int]] = {}

    for col in columns:
        col_values = [row.get(col) for row in rows]
        missing_counts[col] = sum(v is None for v in col_values)
        if inferred_types.get(col) == "numeric":
            numeric_values = [float(v) for v in col_values if isinstance(v, (float, int))]
            if numeric_values:
                numeric_summary[col] = {
                    "min": min(numeric_values),
                    "max": max(numeric_values),
                    "mean": mean(numeric_values),
                }
        else:
            counts = Counter(str(v) for v in col_values if v is not None)
            categorical_summary[col] = dict(counts.most_common(10))

    recommendations = explore_data(rows, inferred_types, numeric_summary, categorical_summary)

    return EDAReport(
        rows=len(rows),
        columns=columns,
        inferred_types=inferred_types,
        missing_counts=missing_counts,
        numeric_summary=numeric_summary,
        categorical_summary=categorical_summary,
        recommendations=recommendations,
    )


def explore_chart(
    rows: list[dict[str, object]],
    inferred_types: dict[str, str],
    output_path: str | Path,
    chart_type: str = "auto",
    x: str | None = None,
    y: str | None = None,
) -> Path:
    """Generate charts (auto/bar/line/hist/box/scatter/pie/area/count/heatmap)."""
    if not rows:
        raise ValueError("Cannot chart an empty dataset")

    numeric_cols = [c for c, t in inferred_types.items() if t == "numeric"]
    categorical_cols = [c for c, t in inferred_types.items() if t == "categorical"]

    if chart_type == "auto":
        chart_type = "bar" if categorical_cols else "hist"

    x_col = x or (categorical_cols[0] if categorical_cols else (numeric_cols[0] if numeric_cols else None))
    y_col = y or (numeric_cols[0] if numeric_cols else None)

    try:
        import matplotlib.pyplot as plt
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("matplotlib is required for chart generation") from exc

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    plt.figure(figsize=(8, 4.5))

    if chart_type in {"bar", "count"}:
        if not x_col:
            raise ValueError("Bar/count chart requires x column")
        counts = Counter(str(r.get(x_col)) for r in rows)
        plt.bar(list(counts.keys()), list(counts.values()))
        plt.xlabel(x_col)
        plt.ylabel("count")

    elif chart_type == "hist":
        col = y_col or x_col
        values = [float(r[col]) for r in rows if isinstance(r.get(col), (float, int))]
        plt.hist(values, bins=10)
        plt.xlabel(col)
        plt.ylabel("frequency")

    elif chart_type == "line":
        if not x_col or not y_col:
            raise ValueError("Line chart requires x and y")
        xs = [str(r.get(x_col)) for r in rows]
        ys = [float(r[y_col]) for r in rows if isinstance(r.get(y_col), (float, int))]
        plt.plot(xs[: len(ys)], ys)
        plt.xlabel(x_col)
        plt.ylabel(y_col)

    elif chart_type == "scatter":
        x_axis = x_col or (numeric_cols[0] if numeric_cols else None)
        y_axis = y_col or (numeric_cols[1] if len(numeric_cols) > 1 else None)
        if not x_axis or not y_axis:
            raise ValueError("Scatter chart requires two numeric columns")
        points = [
            (float(r[x_axis]), float(r[y_axis]))
            for r in rows
            if isinstance(r.get(x_axis), (float, int)) and isinstance(r.get(y_axis), (float, int))
        ]
        plt.scatter([p[0] for p in points], [p[1] for p in points])
        plt.xlabel(x_axis)
        plt.ylabel(y_axis)

    elif chart_type == "box":
        col = y_col or (numeric_cols[0] if numeric_cols else None)
        if not col:
            raise ValueError("Box chart requires numeric column")
        values = [float(r[col]) for r in rows if isinstance(r.get(col), (float, int))]
        plt.boxplot(values)
        plt.ylabel(col)

    elif chart_type == "pie":
        if not x_col:
            raise ValueError("Pie chart requires categorical column")
        counts = Counter(str(r.get(x_col)) for r in rows)
        plt.pie(list(counts.values()), labels=list(counts.keys()), autopct="%1.1f%%")

    elif chart_type == "area":
        col = y_col or (numeric_cols[0] if numeric_cols else None)
        if not col:
            raise ValueError("Area chart requires numeric column")
        values = [float(r[col]) for r in rows if isinstance(r.get(col), (float, int))]
        plt.fill_between(range(len(values)), values)
        plt.ylabel(col)

    elif chart_type == "heatmap":
        import numpy as np

        matrix = []
        for row in rows:
            vals = [float(row[c]) for c in numeric_cols if isinstance(row.get(c), (float, int))]
            if vals:
                matrix.append(vals)
        if not matrix:
            raise ValueError("Heatmap requires numeric data")
        plt.imshow(np.array(matrix).T, aspect="auto", cmap="viridis")
        plt.colorbar()
        plt.yticks(range(len(numeric_cols)), numeric_cols)

    else:
        raise ValueError(f"Unsupported chart_type: {chart_type}")

    plt.title(f"{chart_type.title()} chart")
    plt.tight_layout()
    plt.savefig(out)
    plt.close()
    return out


def generate_bar_chart(rows: list[dict[str, object]], inferred_types: dict[str, str], output_path: str | Path) -> Path:
    """Backwards-compatible convenience wrapper for bar chart generation."""
    return explore_chart(rows, inferred_types, output_path=output_path, chart_type="bar")


def explore_data(
    rows: list[dict[str, object]],
    inferred_types: dict[str, str],
    numeric_summary: dict[str, dict[str, float]] | None = None,
    categorical_summary: dict[str, dict[str, int]] | None = None,
) -> list[str]:
    """Generate insights and next-step analysis recommendations."""
    numeric_summary = numeric_summary or {}
    categorical_summary = categorical_summary or {}

    insights: list[str] = [f"Dataset has {len(rows)} rows and {len(inferred_types)} columns."]

    for col, stats in numeric_summary.items():
        spread = stats["max"] - stats["min"]
        insights.append(f"Numeric column '{col}' has mean={stats['mean']:.3f} and spread={spread:.3f}.")

    for col, counts in categorical_summary.items():
        if counts:
            top, top_count = next(iter(counts.items()))
            insights.append(f"Categorical column '{col}' top value is '{top}' ({top_count} rows).")

    if any(t == "numeric" for t in inferred_types.values()):
        insights.append("Next: consider correlation analysis and regression/classification baselines.")
    if any(t == "categorical" for t in inferred_types.values()):
        insights.append("Next: compare subgroup performance and generate count/bar charts.")
    insights.append("Next: if target exists, run AutoML or prompt user to select supervised/unsupervised analysis.")
    return insights


def _read_rows(path_like: str | Path) -> list[dict[str, object]]:
    path = Path(path_like)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8") as handle:
            return [dict(row) for row in csv.DictReader(handle)]

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:  # pragma: no cover
            raise ModuleNotFoundError("openpyxl is required for Excel support") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(h) for h in next(rows_iter)]
        output: list[dict[str, object]] = []
        for record in rows_iter:
            output.append({headers[i]: record[i] for i in range(len(headers))})
        return output

    raise ValueError(f"Unsupported file type: {path.suffix}. Use CSV or Excel.")


def _is_float(value: str) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False
